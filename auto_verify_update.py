from openpyxl import load_workbook

from cred_sargule import username, password
from tkinter import Tk, Label, Entry, Button
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
import re
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import pygame

excel_file = load_workbook("count.xlsx")
work_sheet = excel_file["Sheet1"]

last_row = work_sheet.max_row
for row in range(last_row, 0, -1):
    if any(cell.value is not None for cell in work_sheet[row]):
        last_non_empty_row = row
        break

current_row = last_non_empty_row + 1


# from captcha import extract_captcha_text
def submit_captcha():
    global captcha_text
    captcha_text = captcha_entry.get()
    window.destroy()


try:

    login_link = "https://rcms.mahafood.gov.in/OfficeLogin.aspx"

    driver = webdriver.Chrome()
    driver.get(login_link)

    time.sleep(3)

    # /html/body/form/div[3]/div[2]/div[2]/div/div/div[3]/div/div[2]/div[2]/div/div/div[2]/input
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$txtUserName").send_keys(
        username
    )

    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$txtPassword").send_keys(
        password
    )

    driver.find_element(
        By.XPATH,
        "/html/body/form/div[3]/div[2]/div[2]/div/div/div[3]/div/div[2]/div[2]/div/div/div[2]/input",
    ).click()
    time.sleep(3)

    captcha_element = driver.find_element(By.ID, "ContentPlaceHolder1_CaptchaImage")
    window = Tk()
    window.geometry("400x150")
    window.title("Captcha Entry")
    window.configure(bg="#f0f8ff")

    captcha_label = Label(
        window, text="Enter Captcha:", font=("Arial", 12), bg="#f0f8ff"
    )
    captcha_label.pack(pady=5)

    captcha_entry = Entry(window, font=("Arial", 12))
    captcha_entry.pack(pady=5)

    # Button to submit captcha and close Tkinter window
    submit_button = Button(
        window,
        text="Validate Captcha",
        font=("Arial", 10, "bold"),
        bg="#4CAF50",  # Green color for the button
        fg="white",
        command=submit_captcha,
    )
    submit_button.pack(pady=10)

    window.mainloop()
    time.sleep(1)
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$txtCaptcha").send_keys(
        captcha_text
    )

    time.sleep(2)

    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$btnUserSignIn").click()

    time.sleep(3)

    TOTP = "985045"  # input("Enter TOTP: ")

    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$Txtusernameuseotp"
    ).send_keys(TOTP)

    time.sleep(5)

    # ctl00$ContentPlaceHolder1$btnUserUseOTP
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$btnUserUseOTP").click()

    # time.sleep(20)

    # driver.find_element(
    #     By.XPATH, "/html/body/div[1]/form/div[3]/div/div/div/div[3]/div/div[2]/div/a"
    # ).click()

    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located(
            (
                By.XPATH,
                "/html/body/div[1]/form/div[3]/div/div/div/div[3]/div/div[2]/div/a",
            )
        )
    )
    element.click()

    time.sleep(3)

    driver.find_element(
        By.XPATH, "/html/body/form/div[3]/div/div[1]/div/div[3]/div/div[1]/ul/li[1]/a"
    ).click()

    time.sleep(3)
    # eKYC Link
    driver.execute_script("window.scrollBy(0, 1000)")

    time.sleep(1)

    driver.find_element(
        By.XPATH,
        "//html/body/form/div[3]/div/div[1]/div/div[3]/div/div[1]/ul/li[1]/ul/li[11]/a",
    ).click()

    time.sleep(3)

    driver.execute_script("window.scrollBy(0, 1000)")

    time.sleep(1)
    # Inside eKYC - 2nd eKYC
    # Link https://rcms.mahafood.gov.in/frmekycdataverification.aspx
    driver.find_element(
        By.XPATH,
        "/html/body/form/div[3]/div/div[1]/div/div[3]/div/div[1]/ul/li[1]/ul/li[11]/ul/li/a",
    ).click()

    time.sleep(5)

    # Reached to the Dashboard

    # /html/body/form/div[3]/div/div[3]/div/div/div[1]/table/tbody/tr[2]/td[21]
    # /html/body/form/div[3]/div/div[3]/div/div/div[1]/table/tbody/tr[21]/td[21]
    # Table >>

    handle = open("count.csv", "a+")

    while True:

        rows = driver.find_elements(
            By.XPATH, "/html/body/form/div[3]/div/div[3]/div/div/div[1]/table/tbody/tr"
        )

        links = []
        links_text = {"Verify and Update": 0, "Reject": 0}
        for row in rows[1:21]:
            cell = row.find_elements(By.TAG_NAME, "td")[20]
            name = row.find_elements(By.TAG_NAME, "td")[2].text
            links.append(
                {"name": name, "link": cell.find_elements(By.TAG_NAME, "a")[0]}
            )  # cell.find_elements(By.TAG_NAME, "a")[0])
            if not cell.find_elements(By.TAG_NAME, "a")[0].text in links_text:
                links_text[cell.find_elements(By.TAG_NAME, "a")[0].text] = 1
            else:
                links_text[cell.find_elements(By.TAG_NAME, "a")[0].text] += 1
        # print(links , "\n")

        if links_text["Verify and Update"] >= 5:
            time.sleep(5)
            driver.find_element(
                By.XPATH,
                "/html/body/form/div[3]/div/div[3]/div/div/div[1]/table/tbody/tr[1]/th[20]/table/tbody/tr[2]/td/input",
            ).click()
            # list_of_checkboxes = driver.find_elements(By.XPATH, "/html/body/form/div[3]/div/div[3]/div/div/div[1]/table/tbody/tr[1]/th[20]/table/tbody/tr[3]/td/input")
            time.sleep(5)
            finalSubmit = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located(
                    (By.NAME, "ctl00$ContentPlaceHolder1$BtnFinalsubmit")
                )
            )
            finalSubmit.click()
            # driver.find_element(
            #     By.NAME, "ctl00$ContentPlaceHolder1$BtnFinalsubmit"
            # ).click()

            alert = WebDriverWait(driver, 5).until(EC.alert_is_present())
            alert.accept()

            alert = WebDriverWait(driver, 100).until(EC.alert_is_present())
            text = str(alert.text)
            match = re.search(r":\s*([0-9]|1[0-9]|20)\s*Unverified", text)

            if match:
                result = match.group(1)
                print("Extracted number:", result)
            else:
                print("No match found")

            time.sleep(2)
            alert.accept()

            time.sleep(2)
            work_sheet[f"A{current_row}"] = str(
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            work_sheet[f"B{current_row}"] = int(result)
            work_sheet[f"C{current_row}"] = f'=TEXT(A{current_row}, "DD/MM/YYYY")'
            current_row += 1
            excel_file.save("count.xlsx")

        else:
            for link in links:
                if link["link"].text == "Reject":
                    print(link["name"], link["link"].get_attribute("href"))
                    link["link"].click()

                    alert = WebDriverWait(driver, 5).until(EC.alert_is_present())
                    alert.accept()

                    time.sleep(5)

                    driver.find_element(
                        By.NAME, "ctl00$ContentPlaceHolder1$Chkdisabilitytype$0"
                    ).click()
                    time.sleep(3)

                    driver.find_element(
                        By.NAME, "ctl00$ContentPlaceHolder1$btnrejectsubmit"
                    ).click()

                    alert = WebDriverWait(driver, 5).until(EC.alert_is_present())
                    alert.accept()

                    alert = WebDriverWait(driver, 100).until(EC.alert_is_present())
                    alert.accept()

                    time.sleep(2)
                    work_sheet[f"A{current_row}"] = str(
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    )
                    work_sheet[f"B{current_row}"] = 1
                    work_sheet[f"C{current_row}"] = (
                        f'=TEXT(A{current_row}, "DD/MM/YYYY")'
                    )
                    current_row += 1
                    excel_file.save("count.xlsx")

                    break
except Exception as e:
    print(e)
    print("Error in code; Playing Music")
    pygame.mixer.init()

    pygame.mixer.music.load("sare_jahan_se_accha.mp3")

    pygame.mixer.music.play()

    while pygame.mixer.music.get_busy():
        pygame.time.Clock().tick(10)
